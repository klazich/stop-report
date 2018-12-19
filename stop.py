import datetime
import os
from os import path
import openpyxl
import click
import helper
import stop_styles as styles
from directories import Directories

dr = Directories()

os.chdir(dr.local.root)

TODAY = datetime.date.today()
SHEET_NAME = TODAY.strftime('%Y%m%d')
SAVE_AS = '{}-WEEK-{}.xlsx'.format(TODAY.strftime('%Y'), TODAY.strftime('%U'))

LOC_SV = path.join(dr.local.stop, SAVE_AS)  # local Excel save path
FDR_SV = path.join(dr.fdrive.stop, SAVE_AS)  # f drive Excel file
ONE_SV = path.join(dr.onedrive.stop, SAVE_AS)  # onedrive Excel file

if path.exists(FDR_SV):
    os.remove(FDR_SV)
if path.exists(ONE_SV):
    os.remove(ONE_SV)


def get_xl_objects():
    """ Returns a workbook and worksheet objects """

    if path.exists(LOC_SV):
        wb = openpyxl.load_workbook(filename=LOC_SV)
        ws = wb.create_sheet('temp')
        if SHEET_NAME in wb.sheetnames:
            wb.remove(wb.get_sheet_by_name(SHEET_NAME))
    else:
        wb = openpyxl.Workbook()
        ws = wb.active

    return wb, ws


def get_axapta_data():
    """ Returns axapta data cleaned and sorted """

    data = helper.get_stop_data()
    data = helper.format_stop_data(data)
    data = helper.sort_stop_data(data)

    return data


def apply_options(ws, data):
    """ Applies page, print and header/footer options from stop_styles.py """

    total = helper.get_value_total(data)

    styles.page_print_options(ws)
    styles.header_footer_options(ws, total=total)
    for k in styles.size_row:
        ws.row_dimensions[int(k)].height = styles.size_row[k]
    for k in styles.size_col:
        ws.column_dimensions[k].width = styles.size_col[k]


def apply_values(ws, data):
    """ Applies values from data to worksheet cells """

    for R, line in enumerate(data):
        for C, value in enumerate(line):
            cell = ws.cell(row=R + 1, column=C + 1)  # get cell object
            styles.apply_to_cell(cell, R, C)
            cell.value = value


def save_reports(wb):
    """ Save workbook to drives """

    wb.save(filename=FDR_SV)
    wb.save(filename=ONE_SV)
    wb.save(filename=LOC_SV)
